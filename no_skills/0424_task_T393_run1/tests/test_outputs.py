import math
import os

import pytest


class TestIrisSpeciesDriverExcel:
    """Tests for verifying iris_species_driver.xlsx output."""

    OUTPUT_PATH = "/root/output/iris_species_driver.xlsx"

    EXPECTED_RESULT = {
        "Top_Measurement": "sepal_width",
        "Contribution": 0.04844315478139366,
        "Avg_R2_Full": 0.4399265499922873,
    }

    EXPECTED_CONTRIBUTIONS = [
        ("sepal_width", 0.04844315478139366),
        ("petal_width", 0.04429056474865212),
        ("petal_length", 0.01085248502159014),
        ("sepal_length", 0.0005440398379050193),
    ]

    TOLERANCE = 1e-6

    def test_file_exists(self):
        """Verify output Excel file was created."""
        assert os.path.exists(self.OUTPUT_PATH), f"Output file not found: {self.OUTPUT_PATH}"

    def test_valid_excel_format(self):
        """Verify output is a valid Excel file that can be loaded."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        assert wb is not None, "Failed to load workbook"

    def test_has_required_sheets(self):
        """Verify workbook contains Result and Contributions sheets."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        assert "Result" in wb.sheetnames, "Missing required sheet: Result"
        assert "Contributions" in wb.sheetnames, "Missing required sheet: Contributions"

    def test_result_sheet_headers(self):
        """Verify Result sheet has correct column headers."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]

        expected_headers = ["Top_Measurement", "Contribution", "Avg_R2_Full"]
        actual_headers = [ws.cell(1, c).value for c in range(1, 4)]

        assert actual_headers == expected_headers, \
            f"Result headers mismatch: expected {expected_headers}, got {actual_headers}"

    def test_result_sheet_top_measurement(self):
        """Verify Top_Measurement value is correct."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]

        actual = ws.cell(2, 1).value
        expected = self.EXPECTED_RESULT["Top_Measurement"]

        assert actual == expected, \
            f"Top_Measurement mismatch: expected {expected}, got {actual}"

    def test_result_sheet_contribution_value(self):
        """Verify Contribution value is correct within tolerance."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]

        actual = ws.cell(2, 2).value
        expected = self.EXPECTED_RESULT["Contribution"]

        assert math.isclose(actual, expected, rel_tol=self.TOLERANCE), \
            f"Contribution mismatch: expected {expected}, got {actual}"

    def test_result_sheet_avg_r2_full_value(self):
        """Verify Avg_R2_Full value is correct within tolerance."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]

        actual = ws.cell(2, 3).value
        expected = self.EXPECTED_RESULT["Avg_R2_Full"]

        assert math.isclose(actual, expected, rel_tol=self.TOLERANCE), \
            f"Avg_R2_Full mismatch: expected {expected}, got {actual}"

    def test_contributions_sheet_headers(self):
        """Verify Contributions sheet has correct column headers."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Contributions"]

        expected_headers = ["Measurement", "Contribution"]
        actual_headers = [ws.cell(1, c).value for c in range(1, 3)]

        assert actual_headers == expected_headers, \
            f"Contributions headers mismatch: expected {expected_headers}, got {actual_headers}"

    def test_contributions_sheet_has_four_measurements(self):
        """Verify Contributions sheet contains all four measurements."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Contributions"]

        measurements = [ws.cell(row, 1).value for row in range(2, 6)]
        expected_measurements = {"sepal_length", "sepal_width", "petal_length", "petal_width"}

        assert set(measurements) == expected_measurements, \
            f"Measurements mismatch: expected {expected_measurements}, got {set(measurements)}"

    def test_contributions_sorted_descending(self):
        """Verify Contributions sheet is sorted by contribution descending."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Contributions"]

        contributions = [ws.cell(row, 2).value for row in range(2, 6)]

        assert contributions == sorted(contributions, reverse=True), \
            f"Contributions not sorted descending: {contributions}"

    def test_contributions_values_correct(self):
        """Verify all contribution values are correct within tolerance."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Contributions"]

        for row_idx, (expected_measurement, expected_contribution) in enumerate(self.EXPECTED_CONTRIBUTIONS, start=2):
            actual_measurement = ws.cell(row_idx, 1).value
            actual_contribution = ws.cell(row_idx, 2).value

            assert actual_measurement == expected_measurement, \
                f"Row {row_idx} measurement mismatch: expected {expected_measurement}, got {actual_measurement}"

            assert math.isclose(actual_contribution, expected_contribution, rel_tol=self.TOLERANCE), \
                f"Row {row_idx} contribution mismatch: expected {expected_contribution}, got {actual_contribution}"

    def test_result_header_formatting_bold(self):
        """Verify Result sheet headers are bold."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]

        for col in range(1, 4):
            cell = ws.cell(1, col)
            assert cell.font.bold, f"Header cell column {col} is not bold"

    def test_result_header_formatting_fill(self):
        """Verify Result sheet headers have light fill."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]

        for col in range(1, 4):
            cell = ws.cell(1, col)
            assert cell.fill.start_color.rgb is not None, \
                f"Header cell column {col} has no fill color"

    def test_result_top_measurement_blue_text(self):
        """Verify Top_Measurement value has blue text (input-style formatting)."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]

        cell = ws.cell(2, 1)
        font_color = cell.font.color.rgb if cell.font.color else None

        assert font_color is not None and "FF" in font_color, \
            f"Top_Measurement should have blue text, got color: {font_color}"

    def test_result_numeric_cells_formatted_4_decimals(self):
        """Verify numeric cells are formatted to 4 decimals."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]

        for col in [2, 3]:
            cell = ws.cell(2, col)
            assert "0.0000" in cell.number_format or cell.number_format == "0.0000", \
                f"Column {col} number format should be 4 decimals, got: {cell.number_format}"

    def test_contributions_sheet_has_chart_image(self):
        """Verify Contributions sheet contains a bar chart image."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Contributions"]

        assert len(ws._images) >= 1, "Contributions sheet should contain at least one image (bar chart)"

    def test_contributions_header_formatting_bold(self):
        """Verify Contributions sheet headers are bold."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Contributions"]

        for col in range(1, 3):
            cell = ws.cell(1, col)
            assert cell.font.bold, f"Contributions header cell column {col} is not bold"

    def test_top_measurement_consistency(self):
        """Verify Top_Measurement in Result matches first row in Contributions."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)

        result_top = wb["Result"].cell(2, 1).value
        contrib_top = wb["Contributions"].cell(2, 1).value

        assert result_top == contrib_top, \
            f"Top measurement inconsistency: Result has {result_top}, Contributions first row has {contrib_top}"

    def test_contribution_value_consistency(self):
        """Verify Contribution in Result matches first contribution in Contributions sheet."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)

        result_contrib = wb["Result"].cell(2, 2).value
        contrib_first = wb["Contributions"].cell(2, 2).value

        assert math.isclose(result_contrib, contrib_first, rel_tol=self.TOLERANCE), \
            f"Contribution value inconsistency: Result has {result_contrib}, Contributions first row has {contrib_first}"
