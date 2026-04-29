import math
import os

import pytest


class TestReconciledSpendVarianceWorkbook:
    """Tests for verifying the reconciled_spend_variance.xlsx workbook."""

    OUTPUT_PATH = "/root/output/reconciled_spend_variance.xlsx"

    EXPECTED_SHEET_NAMES = ["Invoices", "CSV_Totals", "Reconciliation"]

    EXPECTED_INVOICES = [
        {
            "source_file": "invoice.pdf",
            "vendor": "Example, LLC",
            "invoice_number": "123",
            "invoice_date": "March 25, 2024",
            "final_total": 19.00,
        },
        {
            "source_file": "invoice_2.pdf",
            "vendor": "Acme Corporation",
            "invoice_number": "INV-2025-001",
            "invoice_date": "N/A",
            "final_total": 7560.00,
        },
        {
            "source_file": "sample-invoice.pdf",
            "vendor": "Contoso Ltd.",
            "invoice_number": "INV-100",
            "invoice_date": "11/15/2019",
            "final_total": 610.00,
        },
    ]

    EXPECTED_CSV_TOTALS = [
        {"source_file": "purchases.csv", "rule": "SUM(amount)", "computed_total": 5148.35},
        {"source_file": "creditcard.csv", "rule": "SUM(Amount) WHERE Class=1", "computed_total": 60127.97},
        {"source_file": "transactions.csv", "rule": "SUM(col3_amount)", "computed_total": 599232272.78},
    ]

    EXPECTED_INVOICE_TOTAL = 8189.00
    EXPECTED_CSV_SPEND_TOTAL = 599297549.10
    EXPECTED_VARIANCE = 599289360.10

    TOLERANCE = 0.01

    def test_file_exists(self):
        """Verify output Excel file was created."""
        assert os.path.exists(self.OUTPUT_PATH), f"Output file not found at {self.OUTPUT_PATH}"

    def test_valid_xlsx(self):
        """Verify output is a valid Excel workbook."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        assert wb is not None, "Failed to load workbook"

    def test_has_exactly_three_sheets(self):
        """Verify workbook has exactly 3 sheets."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        assert len(wb.sheetnames) == 3, f"Expected 3 sheets, got {len(wb.sheetnames)}"

    def test_sheet_names_correct(self):
        """Verify sheet names are exactly as specified."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        assert wb.sheetnames == self.EXPECTED_SHEET_NAMES, (
            f"Sheet names mismatch: expected {self.EXPECTED_SHEET_NAMES}, got {wb.sheetnames}"
        )


class TestInvoicesSheet:
    """Tests for the Invoices sheet."""

    OUTPUT_PATH = "/root/output/reconciled_spend_variance.xlsx"

    EXPECTED_HEADERS = ["source_file", "vendor", "invoice_number", "invoice_date", "final_total"]

    EXPECTED_INVOICES = [
        {
            "source_file": "invoice.pdf",
            "vendor": "Example, LLC",
            "invoice_number": "123",
            "invoice_date": "March 25, 2024",
            "final_total": 19.00,
        },
        {
            "source_file": "invoice_2.pdf",
            "vendor": "Acme Corporation",
            "invoice_number": "INV-2025-001",
            "invoice_date": "N/A",
            "final_total": 7560.00,
        },
        {
            "source_file": "sample-invoice.pdf",
            "vendor": "Contoso Ltd.",
            "invoice_number": "INV-100",
            "invoice_date": "11/15/2019",
            "final_total": 610.00,
        },
    ]

    TOLERANCE = 0.01

    def test_invoices_sheet_has_correct_headers(self):
        """Verify Invoices sheet has correct column headers."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Invoices"]
        headers = [cell.value for cell in ws[1]]
        assert headers == self.EXPECTED_HEADERS, f"Headers mismatch: expected {self.EXPECTED_HEADERS}, got {headers}"

    def test_invoices_sheet_has_three_data_rows(self):
        """Verify Invoices sheet has exactly 3 invoice rows (plus header)."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Invoices"]
        data_rows = [row for row in ws.iter_rows(min_row=2, max_col=5) if row[0].value is not None]
        assert len(data_rows) == 3, f"Expected 3 invoice rows, got {len(data_rows)}"

    def test_invoice_source_files_correct(self):
        """Verify invoice source file names are correct."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Invoices"]
        for i, expected in enumerate(self.EXPECTED_INVOICES, start=2):
            actual = ws.cell(row=i, column=1).value
            assert actual == expected["source_file"], (
                f"Row {i} source_file mismatch: expected {expected['source_file']}, got {actual}"
            )

    def test_invoice_final_totals_correct(self):
        """Verify invoice final_total values are correct."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Invoices"]
        for i, expected in enumerate(self.EXPECTED_INVOICES, start=2):
            actual = ws.cell(row=i, column=5).value
            assert math.isclose(actual, expected["final_total"], rel_tol=self.TOLERANCE), (
                f"Row {i} final_total mismatch: expected {expected['final_total']}, got {actual}"
            )

    def test_invoice_totals_sum(self):
        """Verify sum of invoice totals is correct."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Invoices"]
        total = sum(ws.cell(row=i, column=5).value for i in range(2, 5))
        expected_sum = sum(inv["final_total"] for inv in self.EXPECTED_INVOICES)
        assert math.isclose(total, expected_sum, rel_tol=self.TOLERANCE), (
            f"Invoice total sum mismatch: expected {expected_sum}, got {total}"
        )


class TestCSVTotalsSheet:
    """Tests for the CSV_Totals sheet."""

    OUTPUT_PATH = "/root/output/reconciled_spend_variance.xlsx"

    EXPECTED_HEADERS = ["source_file", "rule", "computed_total"]

    EXPECTED_CSV_TOTALS = [
        {"source_file": "purchases.csv", "rule": "SUM(amount)", "computed_total": 5148.35},
        {"source_file": "creditcard.csv", "rule": "SUM(Amount) WHERE Class=1", "computed_total": 60127.97},
        {"source_file": "transactions.csv", "rule": "SUM(col3_amount)", "computed_total": 599232272.78},
    ]

    TOLERANCE = 0.01

    def test_csv_totals_sheet_has_correct_headers(self):
        """Verify CSV_Totals sheet has correct column headers."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["CSV_Totals"]
        headers = [cell.value for cell in ws[1]]
        assert headers == self.EXPECTED_HEADERS, f"Headers mismatch: expected {self.EXPECTED_HEADERS}, got {headers}"

    def test_csv_totals_sheet_has_three_data_rows(self):
        """Verify CSV_Totals sheet has exactly 3 data rows (plus header)."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["CSV_Totals"]
        data_rows = [row for row in ws.iter_rows(min_row=2, max_col=3) if row[0].value is not None]
        assert len(data_rows) == 3, f"Expected 3 CSV total rows, got {len(data_rows)}"

    def test_csv_totals_source_files_correct(self):
        """Verify CSV source file names are correct."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["CSV_Totals"]
        for i, expected in enumerate(self.EXPECTED_CSV_TOTALS, start=2):
            actual = ws.cell(row=i, column=1).value
            assert actual == expected["source_file"], (
                f"Row {i} source_file mismatch: expected {expected['source_file']}, got {actual}"
            )

    def test_csv_totals_rules_correct(self):
        """Verify CSV rule descriptions are correct."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["CSV_Totals"]
        for i, expected in enumerate(self.EXPECTED_CSV_TOTALS, start=2):
            actual = ws.cell(row=i, column=2).value
            assert actual == expected["rule"], f"Row {i} rule mismatch: expected {expected['rule']}, got {actual}"

    def test_csv_computed_totals_correct(self):
        """Verify CSV computed_total values are correct."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["CSV_Totals"]
        for i, expected in enumerate(self.EXPECTED_CSV_TOTALS, start=2):
            actual = ws.cell(row=i, column=3).value
            assert math.isclose(actual, expected["computed_total"], rel_tol=self.TOLERANCE), (
                f"Row {i} computed_total mismatch: expected {expected['computed_total']}, got {actual}"
            )


class TestReconciliationSheet:
    """Tests for the Reconciliation sheet."""

    OUTPUT_PATH = "/root/output/reconciled_spend_variance.xlsx"

    EXPECTED_INVOICE_TOTAL = 8189.00
    EXPECTED_CSV_SPEND_TOTAL = 599297549.10
    EXPECTED_VARIANCE = 599289360.10

    TOLERANCE = 0.01

    def test_reconciliation_sheet_has_invoice_total_label(self):
        """Verify INVOICE_TOTAL label exists in cell A2."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Reconciliation"]
        assert ws["A2"].value == "INVOICE_TOTAL", f"A2 should be 'INVOICE_TOTAL', got {ws['A2'].value}"

    def test_reconciliation_sheet_has_csv_spend_total_label(self):
        """Verify CSV_SPEND_TOTAL label exists in cell A4."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Reconciliation"]
        assert ws["A4"].value == "CSV_SPEND_TOTAL", f"A4 should be 'CSV_SPEND_TOTAL', got {ws['A4'].value}"

    def test_reconciliation_sheet_has_variance_label_in_a6(self):
        """Verify VARIANCE label exists in cell A6."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Reconciliation"]
        assert ws["A6"].value == "VARIANCE", f"A6 should be 'VARIANCE', got {ws['A6'].value}"

    def test_invoice_total_formula_correct(self):
        """Verify B2 contains the correct formula for INVOICE_TOTAL."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Reconciliation"]
        formula = ws["B2"].value
        assert formula is not None and "SUM" in str(formula).upper() and "Invoices" in str(formula), (
            f"B2 should contain a SUM formula referencing Invoices sheet, got {formula}"
        )

    def test_csv_spend_total_formula_correct(self):
        """Verify B4 contains the correct formula for CSV_SPEND_TOTAL."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Reconciliation"]
        formula = ws["B4"].value
        assert formula is not None and "SUM" in str(formula).upper() and "CSV_Totals" in str(formula), (
            f"B4 should contain a SUM formula referencing CSV_Totals sheet, got {formula}"
        )

    def test_variance_formula_in_b6(self):
        """Verify B6 contains a formula for VARIANCE (CSV_SPEND_TOTAL - INVOICE_TOTAL)."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Reconciliation"]
        formula = ws["B6"].value
        assert formula is not None and str(formula).startswith("="), (
            f"B6 should contain a formula, got {formula}"
        )
        assert "B4" in str(formula) and "B2" in str(formula), (
            f"B6 formula should reference B4 and B2, got {formula}"
        )

    def test_b2_has_yellow_fill(self):
        """Verify B2 (INVOICE_TOTAL) has yellow fill."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Reconciliation"]
        fill_color = ws["B2"].fill.start_color.rgb if ws["B2"].fill.patternType else None
        assert fill_color == "00FFFF00", f"B2 should have yellow fill (00FFFF00), got {fill_color}"

    def test_b4_has_yellow_fill(self):
        """Verify B4 (CSV_SPEND_TOTAL) has yellow fill."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Reconciliation"]
        fill_color = ws["B4"].fill.start_color.rgb if ws["B4"].fill.patternType else None
        assert fill_color == "00FFFF00", f"B4 should have yellow fill (00FFFF00), got {fill_color}"

    def test_b6_has_yellow_fill(self):
        """Verify B6 (VARIANCE) has yellow fill."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Reconciliation"]
        fill_color = ws["B6"].fill.start_color.rgb if ws["B6"].fill.patternType else None
        assert fill_color == "00FFFF00", f"B6 should have yellow fill (00FFFF00), got {fill_color}"

    def test_currency_format_on_key_cells(self):
        """Verify currency format is applied to key cells."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Reconciliation"]
        expected_format = "$#,##0.00"
        for cell_ref in ["B2", "B4", "B6"]:
            actual_format = ws[cell_ref].number_format
            assert actual_format == expected_format, (
                f"{cell_ref} should have currency format {expected_format}, got {actual_format}"
            )


class TestComputedValues:
    """Tests for verifying the computed values are mathematically correct."""

    OUTPUT_PATH = "/root/output/reconciled_spend_variance.xlsx"

    EXPECTED_INVOICE_TOTAL = 8189.00
    EXPECTED_CSV_SPEND_TOTAL = 599297549.10
    EXPECTED_VARIANCE = 599289360.10

    TOLERANCE = 1.0  # Allow $1 tolerance for large numbers

    def test_invoice_total_value(self):
        """Verify the computed INVOICE_TOTAL matches expected value."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws_inv = wb["Invoices"]
        total = sum(ws_inv.cell(row=i, column=5).value for i in range(2, 5))
        assert math.isclose(total, self.EXPECTED_INVOICE_TOTAL, abs_tol=self.TOLERANCE), (
            f"Invoice total mismatch: expected {self.EXPECTED_INVOICE_TOTAL}, got {total}"
        )

    def test_csv_spend_total_value(self):
        """Verify the computed CSV_SPEND_TOTAL matches expected value."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws_csv = wb["CSV_Totals"]
        total = sum(ws_csv.cell(row=i, column=3).value for i in range(2, 5))
        assert math.isclose(total, self.EXPECTED_CSV_SPEND_TOTAL, abs_tol=self.TOLERANCE), (
            f"CSV spend total mismatch: expected {self.EXPECTED_CSV_SPEND_TOTAL}, got {total}"
        )

    def test_variance_calculation(self):
        """Verify the VARIANCE is CSV_SPEND_TOTAL - INVOICE_TOTAL."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws_inv = wb["Invoices"]
        ws_csv = wb["CSV_Totals"]

        invoice_total = sum(ws_inv.cell(row=i, column=5).value for i in range(2, 5))
        csv_total = sum(ws_csv.cell(row=i, column=3).value for i in range(2, 5))
        computed_variance = csv_total - invoice_total

        assert math.isclose(computed_variance, self.EXPECTED_VARIANCE, abs_tol=self.TOLERANCE), (
            f"Variance mismatch: expected {self.EXPECTED_VARIANCE}, got {computed_variance}"
        )
