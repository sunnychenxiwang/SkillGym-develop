import math
import os

import pytest


class TestSeaLaxWorstCarrierExcel:
    """Tests for verifying the SEA-LAX worst carrier Excel workbook."""

    OUTPUT_PATH = "/root/output/sea_lax_worst_carrier.xlsx"
    TOLERANCE = 0.01

    EXPECTED_SHEET_NAMES = ["Route_Carrier_Summary", "Answer"]
    EXPECTED_HEADERS = ["carrier", "carrier_name", "flight_count_used", "avg_arr_delay_min"]

    EXPECTED_SUMMARY_DATA = [
        {"carrier": "OO", "carrier_name": "SkyWest Airlines Inc.", "flight_count_used": 1020, "avg_arr_delay_min": 7.61},
        {"carrier": "UA", "carrier_name": "United Air Lines Inc.", "flight_count_used": 253, "avg_arr_delay_min": 5.87},
        {"carrier": "AS", "carrier_name": "Alaska Airlines Inc.", "flight_count_used": 4101, "avg_arr_delay_min": 1.92},
        {"carrier": "DL", "carrier_name": "Delta Air Lines Inc.", "flight_count_used": 795, "avg_arr_delay_min": 1.07},
        {"carrier": "VX", "carrier_name": "Virgin America", "flight_count_used": 1261, "avg_arr_delay_min": -2.79},
    ]

    EXPECTED_WORST_CARRIER = {
        "code": "OO",
        "name": "SkyWest Airlines Inc.",
        "avg_arr_delay_min": 7.61,
    }

    def test_file_exists(self):
        """Verify output Excel file was created."""
        assert os.path.exists(self.OUTPUT_PATH), f"Output file not found: {self.OUTPUT_PATH}"

    def test_valid_excel_format(self):
        """Verify output is a valid Excel file that can be opened."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        assert wb is not None, "Failed to load Excel workbook"

    def test_correct_sheet_names(self):
        """Verify workbook contains exactly the required sheets."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        assert wb.sheetnames == self.EXPECTED_SHEET_NAMES, \
            f"Sheet names mismatch: expected {self.EXPECTED_SHEET_NAMES}, got {wb.sheetnames}"

    def test_route_carrier_summary_headers(self):
        """Verify Route_Carrier_Summary has correct headers."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Route_Carrier_Summary"]

        headers = [ws.cell(1, col).value for col in range(1, 5)]
        assert headers == self.EXPECTED_HEADERS, \
            f"Headers mismatch: expected {self.EXPECTED_HEADERS}, got {headers}"

    def test_route_carrier_summary_headers_bold(self):
        """Verify Route_Carrier_Summary header row is bold."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Route_Carrier_Summary"]

        for col in range(1, 5):
            cell = ws.cell(1, col)
            assert cell.font.bold, f"Header '{cell.value}' in column {col} is not bold"

    def test_route_carrier_summary_row_count(self):
        """Verify Route_Carrier_Summary has correct number of data rows."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Route_Carrier_Summary"]

        data_rows = ws.max_row - 1
        expected_rows = len(self.EXPECTED_SUMMARY_DATA)
        assert data_rows == expected_rows, \
            f"Data row count mismatch: expected {expected_rows}, got {data_rows}"

    def test_route_carrier_summary_data_values(self):
        """Verify Route_Carrier_Summary contains correct data values."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Route_Carrier_Summary"]

        for row_idx, expected in enumerate(self.EXPECTED_SUMMARY_DATA, start=2):
            carrier = ws.cell(row_idx, 1).value
            carrier_name = ws.cell(row_idx, 2).value
            flight_count = ws.cell(row_idx, 3).value
            avg_delay = ws.cell(row_idx, 4).value

            assert carrier == expected["carrier"], \
                f"Row {row_idx}: carrier mismatch - expected {expected['carrier']}, got {carrier}"
            assert carrier_name == expected["carrier_name"], \
                f"Row {row_idx}: carrier_name mismatch - expected {expected['carrier_name']}, got {carrier_name}"
            assert flight_count == expected["flight_count_used"], \
                f"Row {row_idx}: flight_count_used mismatch - expected {expected['flight_count_used']}, got {flight_count}"
            assert math.isclose(avg_delay, expected["avg_arr_delay_min"], rel_tol=self.TOLERANCE), \
                f"Row {row_idx}: avg_arr_delay_min mismatch - expected {expected['avg_arr_delay_min']}, got {avg_delay}"

    def test_route_carrier_summary_sorted_correctly(self):
        """Verify Route_Carrier_Summary is sorted by avg_arr_delay_min desc, then carrier asc."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Route_Carrier_Summary"]

        delays = []
        for row in range(2, ws.max_row + 1):
            carrier = ws.cell(row, 1).value
            avg_delay = ws.cell(row, 4).value
            delays.append((avg_delay, carrier))

        for i in range(len(delays) - 1):
            curr_delay, curr_carrier = delays[i]
            next_delay, next_carrier = delays[i + 1]
            if math.isclose(curr_delay, next_delay, rel_tol=self.TOLERANCE):
                assert curr_carrier <= next_carrier, \
                    f"Tie-break sorting error: {curr_carrier} should come before {next_carrier}"
            else:
                assert curr_delay > next_delay, \
                    f"Sorting error: {curr_delay} should be > {next_delay}"

    def test_route_carrier_summary_avg_delay_format(self):
        """Verify avg_arr_delay_min column has 0.00 number format."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Route_Carrier_Summary"]

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, 4)
            assert cell.number_format == "0.00", \
                f"Row {row}: avg_arr_delay_min number format mismatch - expected '0.00', got '{cell.number_format}'"

    def test_answer_sheet_labels(self):
        """Verify Answer sheet has correct labels in column A."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]

        expected_labels = ["worst_carrier_code", "worst_carrier_name", "worst_avg_arr_delay_min"]
        for row, expected_label in enumerate(expected_labels, start=1):
            actual_label = ws[f"A{row}"].value
            assert actual_label == expected_label, \
                f"Row {row}: label mismatch - expected '{expected_label}', got '{actual_label}'"

    def test_answer_sheet_worst_carrier_code(self):
        """Verify Answer sheet has correct worst_carrier_code."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]

        actual = ws["B1"].value
        expected = self.EXPECTED_WORST_CARRIER["code"]
        assert actual == expected, \
            f"worst_carrier_code mismatch: expected '{expected}', got '{actual}'"

    def test_answer_sheet_worst_carrier_name(self):
        """Verify Answer sheet has correct worst_carrier_name."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]

        actual = ws["B2"].value
        expected = self.EXPECTED_WORST_CARRIER["name"]
        assert actual == expected, \
            f"worst_carrier_name mismatch: expected '{expected}', got '{actual}'"

    def test_answer_sheet_cross_sheet_formula(self):
        """Verify Answer sheet B3 contains a cross-sheet Excel formula."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH, data_only=False)
        ws = wb["Answer"]

        formula = ws["B3"].value
        assert isinstance(formula, str), \
            f"B3 should contain a formula string, got {type(formula)}"
        assert formula.startswith("="), \
            f"B3 should be a formula starting with '=', got '{formula}'"
        assert "Route_Carrier_Summary" in formula, \
            f"B3 formula should reference 'Route_Carrier_Summary' sheet, got '{formula}'"

    def test_answer_sheet_worst_delay_format(self):
        """Verify worst_avg_arr_delay_min has 0.00 number format."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]

        cell = ws["B3"]
        assert cell.number_format == "0.00", \
            f"worst_avg_arr_delay_min number format mismatch - expected '0.00', got '{cell.number_format}'"

    def test_answer_sheet_formula_references_correct_cell(self):
        """Verify the cross-sheet formula references the top data row (D2)."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH, data_only=False)
        ws = wb["Answer"]

        formula = ws["B3"].value
        assert "D2" in formula, \
            f"Formula should reference D2 (top data row), got '{formula}'"

    def test_worst_carrier_is_highest_avg_delay(self):
        """Verify the worst carrier has the highest average arrival delay."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws_sum = wb["Route_Carrier_Summary"]
        ws_ans = wb["Answer"]

        first_row_carrier = ws_sum.cell(2, 1).value
        first_row_delay = ws_sum.cell(2, 4).value

        answer_carrier = ws_ans["B1"].value

        assert first_row_carrier == answer_carrier, \
            f"Worst carrier in Answer ({answer_carrier}) should match first row in Summary ({first_row_carrier})"

        max_delay = max(ws_sum.cell(row, 4).value for row in range(2, ws_sum.max_row + 1))
        assert math.isclose(first_row_delay, max_delay, rel_tol=self.TOLERANCE), \
            f"First row delay ({first_row_delay}) should be the maximum ({max_delay})"
