import math
import os

import openpyxl
import pytest


class TestOutputs:
    """Tests for verifying the Excel QC workbook output."""

    OUTPUT_XLSX = "/root/output/read_alignment_qc.xlsx"
    EXPECTED_RESULT = {
        "sheet_names": ["Summary", "Candidates"],
        "summary": {
            "A1": "Source",
            "B1": "learning_bam_file",
            "A2": "best_supported_read_id",
            "B2": "SRR020192.1",
            "A3": "best_supported_read_padj",
            "B3": 0.9736460337665308,
        },
        "candidate_headers": [
            "read_id",
            "origin",
            "sequence_length",
            "gc_fraction",
            "mean_phred",
            "is_aligned",
            "best_mapq",
            "best_ref",
            "best_pos0",
            "deseq2_log2FoldChange",
            "deseq2_padj",
        ],
        "candidates": [
            {
                "read_id": "EAS54_6_R1_2_1_413_324",
                "origin": "example.fastq",
                "sequence_length": 25,
                "gc_fraction": 0.52,
                "mean_phred": 25.28,
                "is_aligned": False,
                "best_mapq": 0,
                "best_ref": None,
                "best_pos0": None,
                "deseq2_log2FoldChange": -0.0777337206075476,
                "deseq2_padj": 0.9736460337665308,
            },
            {
                "read_id": "EAS54_6_R1_2_1_540_792",
                "origin": "example.fastq",
                "sequence_length": 25,
                "gc_fraction": 0.6,
                "mean_phred": 24.52,
                "is_aligned": False,
                "best_mapq": 0,
                "best_ref": None,
                "best_pos0": None,
                "deseq2_log2FoldChange": -0.03371393951576222,
                "deseq2_padj": 0.9736460337665308,
            },
            {
                "read_id": "EAS54_6_R1_2_1_443_348",
                "origin": "example.fastq",
                "sequence_length": 25,
                "gc_fraction": 0.72,
                "mean_phred": 23.4,
                "is_aligned": False,
                "best_mapq": 0,
                "best_ref": None,
                "best_pos0": None,
                "deseq2_log2FoldChange": 0.03370832620089109,
                "deseq2_padj": 0.9736460337665308,
            },
            {
                "read_id": "SRR020192.1",
                "origin": "SRR020192.fastq.gz",
                "sequence_length": 74,
                "gc_fraction": 0.527027027027027,
                "mean_phred": 23.24324324324324,
                "is_aligned": False,
                "best_mapq": 0,
                "best_ref": None,
                "best_pos0": None,
                "deseq2_log2FoldChange": 0.04360253948982735,
                "deseq2_padj": 0.9736460337665308,
            },
        ],
        "freeze_panes": "A2",
        "autofilter": "A1:K5",
        "best_supported_read_id": "SRR020192.1",
        "best_supported_read_padj": 0.9736460337665308,
    }
    TOLERANCE = 1e-9

    def _load_workbook(self):
        return openpyxl.load_workbook(self.OUTPUT_XLSX)

    def _get_candidate_rows(self, worksheet):
        headers = [worksheet.cell(row=1, column=i).value for i in range(1, 12)]
        rows = []
        for row_idx in range(2, worksheet.max_row + 1):
            row = {
                headers[col_idx - 1]: worksheet.cell(row=row_idx, column=col_idx).value
                for col_idx in range(1, 12)
            }
            rows.append(row)
        return rows

    def test_read_alignment_qc_xlsx_exists(self):
        """Verify the required Excel workbook was created."""
        assert os.path.exists(self.OUTPUT_XLSX), f"Output file not found: {self.OUTPUT_XLSX}"

    def test_read_alignment_qc_valid_xlsx(self):
        """Verify the workbook is a valid Excel file readable by openpyxl."""
        workbook = self._load_workbook()
        assert workbook is not None

    def test_workbook_has_exact_required_sheets(self):
        """Verify the workbook contains exactly the required Summary and Candidates sheets."""
        workbook = self._load_workbook()
        assert workbook.sheetnames == self.EXPECTED_RESULT["sheet_names"], (
            f"Sheet names mismatch: expected {self.EXPECTED_RESULT['sheet_names']}, got {workbook.sheetnames}"
        )

    def test_summary_sheet_required_cells(self):
        """Verify the Summary sheet contains the exact required labels and values."""
        workbook = self._load_workbook()
        worksheet = workbook["Summary"]

        for cell_ref, expected_value in self.EXPECTED_RESULT["summary"].items():
            actual_value = worksheet[cell_ref].value
            if isinstance(expected_value, float):
                assert isinstance(actual_value, (int, float)), f"{cell_ref} should contain a number"
                assert math.isclose(actual_value, expected_value, rel_tol=self.TOLERANCE, abs_tol=self.TOLERANCE), (
                    f"{cell_ref} mismatch: expected {expected_value}, got {actual_value}"
                )
            else:
                assert actual_value == expected_value, (
                    f"{cell_ref} mismatch: expected {expected_value!r}, got {actual_value!r}"
                )

    def test_candidates_sheet_headers_match_exact_order(self):
        """Verify the Candidates sheet header row matches the required schema exactly."""
        workbook = self._load_workbook()
        worksheet = workbook["Candidates"]
        actual_headers = [worksheet.cell(row=1, column=i).value for i in range(1, 12)]
        assert actual_headers == self.EXPECTED_RESULT["candidate_headers"], (
            f"Candidate headers mismatch: expected {self.EXPECTED_RESULT['candidate_headers']}, got {actual_headers}"
        )

    def test_candidates_sheet_row_count(self):
        """Verify the Candidates sheet has exactly one row per candidate read."""
        workbook = self._load_workbook()
        worksheet = workbook["Candidates"]
        actual_candidate_count = worksheet.max_row - 1
        expected_candidate_count = len(self.EXPECTED_RESULT["candidates"])
        assert actual_candidate_count == expected_candidate_count, (
            f"Candidate row count mismatch: expected {expected_candidate_count}, got {actual_candidate_count}"
        )

    def test_candidates_values_match_expected(self):
        """Verify each candidate row contains the expected deterministic values."""
        workbook = self._load_workbook()
        worksheet = workbook["Candidates"]
        actual_rows = self._get_candidate_rows(worksheet)

        assert len(actual_rows) == len(self.EXPECTED_RESULT["candidates"])

        for actual_row, expected_row in zip(actual_rows, self.EXPECTED_RESULT["candidates"]):
            for key, expected_value in expected_row.items():
                actual_value = actual_row[key]
                if isinstance(expected_value, float):
                    assert isinstance(actual_value, (int, float)), f"{key} should be numeric for {expected_row['read_id']}"
                    assert math.isclose(actual_value, expected_value, rel_tol=self.TOLERANCE, abs_tol=self.TOLERANCE), (
                        f"{key} mismatch for {expected_row['read_id']}: expected {expected_value}, got {actual_value}"
                    )
                else:
                    assert actual_value == expected_value, (
                        f"{key} mismatch for {expected_row['read_id']}: expected {expected_value!r}, got {actual_value!r}"
                    )

    def test_candidates_required_fields_and_types(self):
        """Verify all candidate rows contain required fields with valid data types and ranges."""
        workbook = self._load_workbook()
        worksheet = workbook["Candidates"]
        rows = self._get_candidate_rows(worksheet)
        required_fields = self.EXPECTED_RESULT["candidate_headers"]

        for row in rows:
            assert set(row.keys()) == set(required_fields), f"Candidate row keys mismatch for {row['read_id']}"
            assert isinstance(row["read_id"], str) and row["read_id"], "read_id must be a non-empty string"
            assert row["origin"] in {"example.fastq", "SRR020192.fastq.gz"}, "origin must be one of the allowed file names"
            assert isinstance(row["sequence_length"], int) and row["sequence_length"] > 0, "sequence_length must be a positive integer"
            assert isinstance(row["gc_fraction"], (int, float)), "gc_fraction must be numeric"
            assert 0.0 <= row["gc_fraction"] <= 1.0, "gc_fraction must be between 0 and 1"
            assert isinstance(row["mean_phred"], (int, float)), "mean_phred must be numeric"
            assert row["mean_phred"] >= 0.0, "mean_phred must be non-negative"
            assert isinstance(row["is_aligned"], bool), "is_aligned must be boolean"
            assert isinstance(row["best_mapq"], int), "best_mapq must be an integer"
            assert row["best_mapq"] >= 0, "best_mapq must be non-negative"
            assert row["best_ref"] is None or isinstance(row["best_ref"], str), "best_ref must be blank or a string"
            assert row["best_pos0"] is None or isinstance(row["best_pos0"], int), "best_pos0 must be blank or an integer"
            assert isinstance(row["deseq2_log2FoldChange"], (int, float)), "deseq2_log2FoldChange must be numeric"
            assert isinstance(row["deseq2_padj"], (int, float)), "deseq2_padj must be numeric"
            assert 0.0 <= row["deseq2_padj"] <= 1.0, "deseq2_padj must be between 0 and 1"

    def test_unaligned_reads_have_blank_reference_fields(self):
        """Verify unaligned reads have zero MAPQ and blank reference/position fields."""
        workbook = self._load_workbook()
        worksheet = workbook["Candidates"]
        rows = self._get_candidate_rows(worksheet)

        for row in rows:
            if row["is_aligned"] is False:
                assert row["best_mapq"] == 0, f"Unaligned read {row['read_id']} should have best_mapq 0"
                assert row["best_ref"] is None, f"Unaligned read {row['read_id']} should have blank best_ref"
                assert row["best_pos0"] is None, f"Unaligned read {row['read_id']} should have blank best_pos0"

    def test_candidates_formatting(self):
        """Verify the Candidates sheet has the required freeze pane, bold header, and autofilter."""
        workbook = self._load_workbook()
        worksheet = workbook["Candidates"]

        assert worksheet.freeze_panes == self.EXPECTED_RESULT["freeze_panes"], (
            f"Freeze panes mismatch: expected {self.EXPECTED_RESULT['freeze_panes']}, got {worksheet.freeze_panes}"
        )
        assert worksheet.auto_filter.ref == self.EXPECTED_RESULT["autofilter"], (
            f"Autofilter mismatch: expected {self.EXPECTED_RESULT['autofilter']}, got {worksheet.auto_filter.ref}"
        )
        for col_idx in range(1, 12):
            assert worksheet.cell(row=1, column=col_idx).font.bold is True, (
                f"Header cell {worksheet.cell(row=1, column=col_idx).coordinate} should be bold"
            )

    def test_best_supported_read_matches_summary_and_candidates(self):
        """Verify the Summary best-supported read exists in Candidates and matches its padj value."""
        workbook = self._load_workbook()
        summary_ws = workbook["Summary"]
        candidates_ws = workbook["Candidates"]
        rows = self._get_candidate_rows(candidates_ws)

        best_read_id = summary_ws["B2"].value
        best_padj = summary_ws["B3"].value

        matching_rows = [row for row in rows if row["read_id"] == best_read_id]
        assert len(matching_rows) == 1, f"Expected exactly one candidate row for best-supported read {best_read_id}"

        candidate_row = matching_rows[0]
        assert math.isclose(candidate_row["deseq2_padj"], best_padj, rel_tol=self.TOLERANCE, abs_tol=self.TOLERANCE), (
            f"Summary padj {best_padj} does not match candidate row padj {candidate_row['deseq2_padj']}"
        )

    def test_best_supported_read_is_uniquely_selected_by_rule(self):
        """Verify the Summary best-supported read is the unique winner by padj, then log2FoldChange, then read_id."""
        workbook = self._load_workbook()
        summary_ws = workbook["Summary"]
        candidates_ws = workbook["Candidates"]
        rows = self._get_candidate_rows(candidates_ws)

        def sort_key(row):
            return (row["deseq2_padj"], -row["deseq2_log2FoldChange"], row["read_id"])

        best_row = sorted(rows, key=sort_key)[0]
        assert summary_ws["B2"].value == best_row["read_id"], (
            f"Expected best-supported read {best_row['read_id']}, got {summary_ws['B2'].value}"
        )
        assert best_row["read_id"] == self.EXPECTED_RESULT["best_supported_read_id"]
        assert math.isclose(best_row["deseq2_padj"], self.EXPECTED_RESULT["best_supported_read_padj"], rel_tol=self.TOLERANCE, abs_tol=self.TOLERANCE)

    @pytest.mark.parametrize(
        "read_id,expected_origin",
        [
            ("EAS54_6_R1_2_1_413_324", "example.fastq"),
            ("EAS54_6_R1_2_1_540_792", "example.fastq"),
            ("EAS54_6_R1_2_1_443_348", "example.fastq"),
            ("SRR020192.1", "SRR020192.fastq.gz"),
        ],
    )
    def test_candidate_origins(self, read_id, expected_origin):
        """Verify each expected read ID is present with the correct source origin."""
        workbook = self._load_workbook()
        worksheet = workbook["Candidates"]
        rows = self._get_candidate_rows(worksheet)
        row_map = {row["read_id"]: row for row in rows}

        assert read_id in row_map, f"Missing candidate read_id: {read_id}"
        assert row_map[read_id]["origin"] == expected_origin, (
            f"Origin mismatch for {read_id}: expected {expected_origin}, got {row_map[read_id]['origin']}"
        )
