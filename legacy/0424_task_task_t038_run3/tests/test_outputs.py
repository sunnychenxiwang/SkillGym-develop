import math
import os

import pytest


class TestBlankDistortionExcel:
    """Tests for verifying the blank_distortion_construction_scaled.xlsx output."""

    OUTPUT_PATH = "/root/output/blank_distortion_construction_scaled.xlsx"

    EXPECTED_RESULT = {
        "BDR": 0.05263157894736842,
        "Permits_Total_GrandMean": 1223.0,
        "Spending_CurrentCombinedTotal_GrandMean": 80853.33727810651,
    }
    TOLERANCE = 1e-6

    # Structural Tests

    def test_file_exists(self):
        """Verify output Excel file was created."""
        assert os.path.exists(self.OUTPUT_PATH), f"Output file not found: {self.OUTPUT_PATH}"

    def test_valid_xlsx(self):
        """Verify output is a valid Excel file that can be loaded."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        assert wb is not None, "Failed to load workbook"

    def test_has_result_sheet(self):
        """Verify Result sheet exists."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        assert "Result" in wb.sheetnames, "Missing required sheet: Result"

    def test_has_scaled_impact_sheet(self):
        """Verify Scaled_Impact sheet exists."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        assert "Scaled_Impact" in wb.sheetnames, "Missing required sheet: Scaled_Impact"

    # Value Tests - Result Sheet

    def test_bdr_value(self):
        """Verify BDR value is correct."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        bdr_value = ws["B1"].value
        assert math.isclose(bdr_value, self.EXPECTED_RESULT["BDR"], rel_tol=self.TOLERANCE), \
            f"BDR mismatch: expected {self.EXPECTED_RESULT['BDR']}, got {bdr_value}"

    def test_permits_grand_mean_value(self):
        """Verify Permits_Total_GrandMean value is correct."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        permits_value = ws["B2"].value
        assert math.isclose(permits_value, self.EXPECTED_RESULT["Permits_Total_GrandMean"], rel_tol=self.TOLERANCE), \
            f"Permits_Total_GrandMean mismatch: expected {self.EXPECTED_RESULT['Permits_Total_GrandMean']}, got {permits_value}"

    def test_spending_grand_mean_value(self):
        """Verify Spending_CurrentCombinedTotal_GrandMean value is correct."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        spending_value = ws["B3"].value
        assert math.isclose(spending_value, self.EXPECTED_RESULT["Spending_CurrentCombinedTotal_GrandMean"], rel_tol=self.TOLERANCE), \
            f"Spending mismatch: expected {self.EXPECTED_RESULT['Spending_CurrentCombinedTotal_GrandMean']}, got {spending_value}"

    # Content Tests - Result Sheet Labels

    def test_result_sheet_bdr_label(self):
        """Verify BDR label in Result sheet."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        assert ws["A1"].value == "BDR", f"Expected label 'BDR', got '{ws['A1'].value}'"

    def test_result_sheet_permits_label(self):
        """Verify Permits_Total_GrandMean label in Result sheet."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        assert ws["A2"].value == "Permits_Total_GrandMean", \
            f"Expected label 'Permits_Total_GrandMean', got '{ws['A2'].value}'"

    def test_result_sheet_spending_label(self):
        """Verify Spending_CurrentCombinedTotal_GrandMean label in Result sheet."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        assert ws["A3"].value == "Spending_CurrentCombinedTotal_GrandMean", \
            f"Expected label 'Spending_CurrentCombinedTotal_GrandMean', got '{ws['A3'].value}'"

    # Content Tests - Result Sheet Formatting

    def test_result_labels_bold(self):
        """Verify all labels in Result sheet are bold."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        assert ws["A1"].font.bold, "BDR label should be bold"
        assert ws["A2"].font.bold, "Permits_Total_GrandMean label should be bold"
        assert ws["A3"].font.bold, "Spending_CurrentCombinedTotal_GrandMean label should be bold"

    def test_result_values_blue_font(self):
        """Verify numeric values in Result sheet have blue font."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]

        for cell_ref in ["B1", "B2", "B3"]:
            cell = ws[cell_ref]
            color = cell.font.color.rgb if cell.font.color else None
            assert color == "000000FF", f"Cell {cell_ref} should have blue font (000000FF), got {color}"

    def test_bdr_number_format(self):
        """Verify BDR has 6 decimal format."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        assert ws["B1"].number_format == "0.000000", \
            f"BDR format should be '0.000000', got '{ws['B1'].number_format}'"

    def test_permits_number_format(self):
        """Verify Permits_Total_GrandMean has #,##0.00 format."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        assert ws["B2"].number_format == "#,##0.00", \
            f"Permits format should be '#,##0.00', got '{ws['B2'].number_format}'"

    def test_spending_number_format(self):
        """Verify Spending_CurrentCombinedTotal_GrandMean has $#,##0.00 format."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        assert ws["B3"].number_format == "$#,##0.00", \
            f"Spending format should be '$#,##0.00', got '{ws['B3'].number_format}'"

    # Content Tests - Scaled_Impact Sheet

    def test_scaled_impact_permits_label(self):
        """Verify BDR * Permits_Total_GrandMean label in Scaled_Impact sheet."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Scaled_Impact"]
        assert ws["A1"].value == "BDR * Permits_Total_GrandMean", \
            f"Expected label 'BDR * Permits_Total_GrandMean', got '{ws['A1'].value}'"

    def test_scaled_impact_spending_label(self):
        """Verify BDR * Spending_CurrentCombinedTotal_GrandMean label in Scaled_Impact sheet."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Scaled_Impact"]
        assert ws["A2"].value == "BDR * Spending_CurrentCombinedTotal_GrandMean", \
            f"Expected label 'BDR * Spending_CurrentCombinedTotal_GrandMean', got '{ws['A2'].value}'"

    def test_scaled_impact_permits_formula(self):
        """Verify BDR * Permits formula in Scaled_Impact sheet."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Scaled_Impact"]
        formula = ws["B1"].value
        assert formula == "=Result!B1*Result!B2", \
            f"Expected formula '=Result!B1*Result!B2', got '{formula}'"

    def test_scaled_impact_spending_formula(self):
        """Verify BDR * Spending formula in Scaled_Impact sheet."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Scaled_Impact"]
        formula = ws["B2"].value
        assert formula == "=Result!B1*Result!B3", \
            f"Expected formula '=Result!B1*Result!B3', got '{formula}'"

    def test_scaled_impact_black_font(self):
        """Verify formula cells in Scaled_Impact sheet have black font."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Scaled_Impact"]

        for cell_ref in ["B1", "B2"]:
            cell = ws[cell_ref]
            color = cell.font.color.rgb if cell.font.color else None
            assert color == "00000000", f"Cell {cell_ref} should have black font (00000000), got {color}"

    def test_scaled_impact_permits_format(self):
        """Verify BDR * Permits formula cell has #,##0.00 format."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Scaled_Impact"]
        assert ws["B1"].number_format == "#,##0.00", \
            f"Scaled permits format should be '#,##0.00', got '{ws['B1'].number_format}'"

    def test_scaled_impact_spending_format(self):
        """Verify BDR * Spending formula cell has $#,##0.00 format."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Scaled_Impact"]
        assert ws["B2"].number_format == "$#,##0.00", \
            f"Scaled spending format should be '$#,##0.00', got '{ws['B2'].number_format}'"

    # Data Integrity Tests

    def test_bdr_reasonable_range(self):
        """Verify BDR is in a reasonable range (0 to 1 for a ratio)."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        bdr_value = ws["B1"].value
        assert 0 <= bdr_value <= 1, f"BDR should be between 0 and 1, got {bdr_value}"

    def test_permits_positive(self):
        """Verify Permits_Total_GrandMean is positive."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        permits_value = ws["B2"].value
        assert permits_value > 0, f"Permits_Total_GrandMean should be positive, got {permits_value}"

    def test_spending_positive(self):
        """Verify Spending_CurrentCombinedTotal_GrandMean is positive."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]
        spending_value = ws["B3"].value
        assert spending_value > 0, f"Spending should be positive, got {spending_value}"

    def test_values_numeric(self):
        """Verify all values in Result sheet B column are numeric."""
        from openpyxl import load_workbook
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Result"]

        for row in range(1, 4):
            value = ws[f"B{row}"].value
            assert isinstance(value, (int, float)), \
                f"Cell B{row} should be numeric, got {type(value).__name__}"
