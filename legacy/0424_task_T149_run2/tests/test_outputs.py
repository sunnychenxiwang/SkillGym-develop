import math
import os

import pytest


class TestDelayVsDemandExcel:
    """Tests for verifying the delay_vs_demand.xlsx output file."""

    OUTPUT_PATH = "/root/output/delay_vs_demand.xlsx"

    EXPECTED_SHEETS = ["monthly_joined", "result"]
    EXPECTED_MONTHLY_HEADERS = ["year", "month", "monthly_passengers", "avg_arr_delay"]
    EXPECTED_RESULT_A1 = "correlation"
    EXPECTED_CORRELATION_FORMAT = "0.0000"

    # Tolerance for float comparisons
    TOLERANCE = 0.0001

    def test_file_exists(self):
        """Verify output Excel file was created at the specified path."""
        assert os.path.exists(self.OUTPUT_PATH), \
            f"Output file not found at {self.OUTPUT_PATH}"

    def test_valid_excel_format(self):
        """Verify output is a valid Excel file that can be opened."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        assert wb is not None, "Failed to load Excel workbook"
        wb.close()

    def test_has_required_sheets(self):
        """Verify Excel file has both required sheets: monthly_joined and result."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)

        for sheet_name in self.EXPECTED_SHEETS:
            assert sheet_name in wb.sheetnames, \
                f"Missing required sheet: {sheet_name}. Found sheets: {wb.sheetnames}"

        wb.close()

    def test_monthly_joined_has_correct_headers(self):
        """Verify monthly_joined sheet has exactly the required column headers."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["monthly_joined"]

        actual_headers = [ws.cell(1, col).value for col in range(1, 5)]

        assert actual_headers == self.EXPECTED_MONTHLY_HEADERS, \
            f"Headers mismatch. Expected: {self.EXPECTED_MONTHLY_HEADERS}, Got: {actual_headers}"

        wb.close()

    def test_monthly_joined_headers_are_bold(self):
        """Verify all headers in monthly_joined sheet are bold formatted."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["monthly_joined"]

        for col in range(1, 5):
            cell = ws.cell(1, col)
            assert cell.font.bold, \
                f"Header '{cell.value}' at column {col} is not bold"

        wb.close()

    def test_result_sheet_a1_is_correlation_label(self):
        """Verify result sheet has 'correlation' label in cell A1."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["result"]

        assert ws["A1"].value == self.EXPECTED_RESULT_A1, \
            f"A1 should be '{self.EXPECTED_RESULT_A1}', got '{ws['A1'].value}'"

        wb.close()

    def test_result_sheet_a1_is_bold(self):
        """Verify 'correlation' label in A1 is bold formatted."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["result"]

        assert ws["A1"].font.bold, "A1 'correlation' label should be bold"

        wb.close()

    def test_result_sheet_b1_has_correlation_value(self):
        """Verify result sheet has correlation value in cell B1."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["result"]

        b1_value = ws["B1"].value

        # B1 should contain a numeric correlation value or None/NaN if no data
        # Given the data doesn't overlap, NaN (stored as None) is acceptable
        is_numeric = isinstance(b1_value, (int, float))
        is_nan_or_none = b1_value is None or (isinstance(b1_value, float) and math.isnan(b1_value))

        assert is_numeric or is_nan_or_none, \
            f"B1 should contain a numeric correlation value or None/NaN, got: {type(b1_value).__name__} = {b1_value}"

        wb.close()

    def test_result_sheet_b1_number_format(self):
        """Verify correlation value in B1 has 4 decimal number format."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["result"]

        assert ws["B1"].number_format == self.EXPECTED_CORRELATION_FORMAT, \
            f"B1 number format should be '{self.EXPECTED_CORRELATION_FORMAT}', got '{ws['B1'].number_format}'"

        wb.close()

    def test_result_sheet_b1_font_color_is_black(self):
        """Verify correlation value cell B1 has black font color."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["result"]

        font_color = ws["B1"].font.color

        # Black color can be represented as rgb='00000000' or theme color
        if font_color.rgb:
            # RGB format - should be black (00000000 or 000000)
            rgb_value = font_color.rgb.lower()
            is_black = rgb_value in ("00000000", "ff000000", "000000")
            assert is_black, \
                f"B1 font color should be black, got RGB: {font_color.rgb}"

        wb.close()

    def test_monthly_joined_column_count(self):
        """Verify monthly_joined sheet has exactly 4 columns."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["monthly_joined"]

        # Check that columns 1-4 have headers and column 5 is empty
        headers = [ws.cell(1, col).value for col in range(1, 6)]
        non_empty_headers = [h for h in headers[:4] if h is not None]

        assert len(non_empty_headers) == 4, \
            f"Expected 4 columns, found headers: {headers}"
        assert headers[4] is None, \
            f"Column 5 should be empty, found: {headers[4]}"

        wb.close()

    def test_monthly_joined_header_order(self):
        """Verify monthly_joined headers are in exact required order."""
        from openpyxl import load_workbook

        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["monthly_joined"]

        expected_order = ["year", "month", "monthly_passengers", "avg_arr_delay"]
        actual_order = [ws.cell(1, col).value for col in range(1, 5)]

        for i, (expected, actual) in enumerate(zip(expected_order, actual_order)):
            assert expected == actual, \
                f"Column {i+1} header should be '{expected}', got '{actual}'"

        wb.close()


class TestExcelFileIntegrity:
    """Tests for verifying Excel file integrity and structure."""

    OUTPUT_PATH = "/root/output/delay_vs_demand.xlsx"

    def test_file_is_valid_zip(self):
        """Verify Excel file is a valid ZIP archive (OOXML format)."""
        import zipfile

        assert zipfile.is_zipfile(self.OUTPUT_PATH), \
            "Excel file should be a valid ZIP archive (OOXML format)"

    def test_file_not_empty(self):
        """Verify output file is not empty."""
        file_size = os.path.getsize(self.OUTPUT_PATH)
        assert file_size > 0, "Output file should not be empty"

    def test_file_has_reasonable_size(self):
        """Verify output file has a reasonable size for an Excel workbook."""
        file_size = os.path.getsize(self.OUTPUT_PATH)
        # Excel file should be at least 1KB and less than 10MB for this task
        assert file_size > 1024, f"File seems too small: {file_size} bytes"
        assert file_size < 10 * 1024 * 1024, f"File seems too large: {file_size} bytes"

    def test_contains_expected_internal_structure(self):
        """Verify Excel file contains expected internal XML structure."""
        import zipfile

        with zipfile.ZipFile(self.OUTPUT_PATH, 'r') as zf:
            file_list = zf.namelist()

            # Check for essential OOXML components
            assert any('[Content_Types].xml' in f for f in file_list), \
                "Missing [Content_Types].xml"
            assert any('workbook.xml' in f for f in file_list), \
                "Missing workbook.xml"
            assert any('sheet' in f.lower() for f in file_list), \
                "Missing sheet files"
