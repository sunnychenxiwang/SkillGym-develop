import math
import os

import pandas as pd
import pytest
from openpyxl import load_workbook


class TestGrowthVsConstructionCorrelation:
    """Tests for verifying the growth_vs_construction_correlation.xlsx output."""

    OUTPUT_PATH = "/root/output/growth_vs_construction_correlation.xlsx"

    EXPECTED_SECTORS = [
        "annual.combined.residential",
        "annual.combined.commercial",
        "annual.combined.educational",
    ]

    EXPECTED_CORRELATIONS = {
        "annual.combined.residential": -0.449129,
        "annual.combined.commercial": -0.056356,
        "annual.combined.educational": 0.795309,
    }

    EXPECTED_AVG_TOTAL_PERMITS = 1223.00
    EXPECTED_BEST_MATCH_SECTOR = "annual.combined.educational"
    EXPECTED_DATA_USED_ROWS = 120

    TOLERANCE = 1e-5

    # --- Structural Tests ---

    def test_file_exists(self):
        """Verify output Excel file was created at the expected path."""
        assert os.path.exists(self.OUTPUT_PATH), f"Output file not found: {self.OUTPUT_PATH}"

    def test_valid_xlsx(self):
        """Verify output is a valid Excel file that can be loaded."""
        wb = load_workbook(self.OUTPUT_PATH)
        assert wb is not None, "Failed to load Excel workbook"
        wb.close()

    def test_results_sheet_exists(self):
        """Verify Results sheet exists in workbook."""
        wb = load_workbook(self.OUTPUT_PATH)
        assert "Results" in wb.sheetnames, "Missing required sheet: Results"
        wb.close()

    def test_data_used_sheet_exists(self):
        """Verify Data_Used sheet exists in workbook."""
        wb = load_workbook(self.OUTPUT_PATH)
        assert "Data_Used" in wb.sheetnames, "Missing required sheet: Data_Used"
        wb.close()

    # --- Results Sheet Content Tests ---

    def test_results_has_required_columns(self):
        """Verify Results sheet has all required columns."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Results")
        required_columns = ["sector_column", "corr_with_OD_mean", "avg_total_permits", "best_match"]
        for col in required_columns:
            assert col in df.columns, f"Missing required column in Results: {col}"

    def test_results_has_three_rows(self):
        """Verify Results sheet has exactly 3 data rows (one per sector)."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Results")
        assert len(df) == 3, f"Expected 3 rows in Results sheet, got {len(df)}"

    def test_results_sector_columns_correct(self):
        """Verify Results sheet contains the correct sector column names."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Results")
        actual_sectors = df["sector_column"].tolist()
        for sector in self.EXPECTED_SECTORS:
            assert sector in actual_sectors, f"Missing sector: {sector}"

    def test_results_correlations_correct(self):
        """Verify correlation values are calculated correctly."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Results")
        for _, row in df.iterrows():
            sector = row["sector_column"]
            actual_corr = row["corr_with_OD_mean"]
            expected_corr = self.EXPECTED_CORRELATIONS[sector]
            assert math.isclose(actual_corr, expected_corr, rel_tol=self.TOLERANCE), \
                f"Correlation mismatch for {sector}: expected {expected_corr}, got {actual_corr}"

    def test_results_avg_total_permits_correct(self):
        """Verify avg_total_permits is the same for all rows and matches expected value."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Results")
        for _, row in df.iterrows():
            actual_permits = row["avg_total_permits"]
            assert math.isclose(actual_permits, self.EXPECTED_AVG_TOTAL_PERMITS, rel_tol=0.01), \
                f"avg_total_permits mismatch: expected {self.EXPECTED_AVG_TOTAL_PERMITS}, got {actual_permits}"

    def test_results_best_match_only_one_true(self):
        """Verify exactly one row has best_match = TRUE."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Results")
        true_count = df["best_match"].sum()
        assert true_count == 1, f"Expected exactly 1 best_match=TRUE, got {true_count}"

    def test_results_best_match_is_educational(self):
        """Verify the best_match=TRUE row is for the educational sector (highest correlation)."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Results")
        best_row = df[df["best_match"] == True]
        assert len(best_row) == 1, "Expected exactly one best_match row"
        actual_best_sector = best_row.iloc[0]["sector_column"]
        assert actual_best_sector == self.EXPECTED_BEST_MATCH_SECTOR, \
            f"best_match sector mismatch: expected {self.EXPECTED_BEST_MATCH_SECTOR}, got {actual_best_sector}"

    def test_results_best_match_has_highest_correlation(self):
        """Verify the best_match row has the highest correlation value."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Results")
        best_row = df[df["best_match"] == True]
        best_corr = best_row.iloc[0]["corr_with_OD_mean"]
        max_corr = df["corr_with_OD_mean"].max()
        assert math.isclose(best_corr, max_corr, rel_tol=self.TOLERANCE), \
            f"best_match row does not have highest correlation: best={best_corr}, max={max_corr}"

    # --- Data_Used Sheet Content Tests ---

    def test_data_used_has_required_columns(self):
        """Verify Data_Used sheet has required columns."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Data_Used")
        required_columns = ["OD_mean", "best_sector_series"]
        for col in required_columns:
            assert col in df.columns, f"Missing required column in Data_Used: {col}"

    def test_data_used_has_expected_rows(self):
        """Verify Data_Used sheet has the expected number of rows (alignment length N)."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Data_Used")
        assert len(df) == self.EXPECTED_DATA_USED_ROWS, \
            f"Expected {self.EXPECTED_DATA_USED_ROWS} rows in Data_Used, got {len(df)}"

    def test_data_used_od_mean_values_numeric(self):
        """Verify OD_mean column contains numeric values."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Data_Used")
        assert pd.api.types.is_numeric_dtype(df["OD_mean"]), "OD_mean column should be numeric"

    def test_data_used_best_sector_series_values_numeric(self):
        """Verify best_sector_series column contains numeric values."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Data_Used")
        assert pd.api.types.is_numeric_dtype(df["best_sector_series"]), \
            "best_sector_series column should be numeric"

    def test_data_used_no_missing_values(self):
        """Verify Data_Used sheet has no missing values."""
        df = pd.read_excel(self.OUTPUT_PATH, sheet_name="Data_Used")
        assert df.isna().sum().sum() == 0, "Data_Used sheet should have no missing values"

    # --- Formatting Tests ---

    def test_results_header_bold(self):
        """Verify Results sheet header row is bold."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Results"]
        for cell in ws[1]:
            assert cell.font.bold, f"Header cell {cell.coordinate} should be bold"
        wb.close()

    def test_results_header_has_fill(self):
        """Verify Results sheet header row has a light fill color."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Results"]
        for cell in ws[1]:
            fill_color = ws[cell.coordinate].fill.fgColor.rgb
            assert fill_color is not None and fill_color != "00000000", \
                f"Header cell {cell.coordinate} should have a fill color"
        wb.close()

    def test_results_freeze_top_row(self):
        """Verify Results sheet has frozen top row."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Results"]
        assert ws.freeze_panes == "A2", f"Expected freeze_panes='A2', got {ws.freeze_panes}"
        wb.close()

    def test_results_autofilter_enabled(self):
        """Verify Results sheet has autofilter enabled."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Results"]
        assert ws.auto_filter.ref is not None, "Results sheet should have autofilter enabled"
        wb.close()

    def test_results_correlation_number_format(self):
        """Verify corr_with_OD_mean column uses 0.000000 format."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Results"]
        headers = [c.value for c in ws[1]]
        corr_col = headers.index("corr_with_OD_mean") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, corr_col)
            assert cell.number_format == "0.000000", \
                f"Cell {cell.coordinate} should have format '0.000000', got '{cell.number_format}'"
        wb.close()

    def test_results_permits_number_format(self):
        """Verify avg_total_permits column uses 0.00 format."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Results"]
        headers = [c.value for c in ws[1]]
        perm_col = headers.index("avg_total_permits") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, perm_col)
            assert cell.number_format == "0.00", \
                f"Cell {cell.coordinate} should have format '0.00', got '{cell.number_format}'"
        wb.close()

    def test_data_used_header_bold(self):
        """Verify Data_Used sheet header row is bold."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Data_Used"]
        for cell in ws[1]:
            assert cell.font.bold, f"Header cell {cell.coordinate} should be bold"
        wb.close()

    def test_data_used_freeze_top_row(self):
        """Verify Data_Used sheet has frozen top row."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Data_Used"]
        assert ws.freeze_panes == "A2", f"Expected freeze_panes='A2', got {ws.freeze_panes}"
        wb.close()

    def test_data_used_autofilter_enabled(self):
        """Verify Data_Used sheet has autofilter enabled."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Data_Used"]
        assert ws.auto_filter.ref is not None, "Data_Used sheet should have autofilter enabled"
        wb.close()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
