"""
Tests for verifying the spending_vs_growth_answer.xlsx output file.
"""

import math
import os

import pytest
from openpyxl import load_workbook


class TestSpendingVsGrowthWorkbook:
    """Tests for the Excel workbook output."""

    OUTPUT_PATH = "/root/output/spending_vs_growth_answer.xlsx"

    EXPECTED_RESULT = {
        "mean_max_growth_rate": 0.3337191358024701,
        "best_year": 2007,
        "annual_total_spending": 35352893,
        "spending_intensity": 105936067.8104043,
        "avg_monthly_total": None,  # blank because 2007 not in construction.csv
    }

    EXPECTED_HEADERS = [
        "mean_max_growth_rate",
        "best_year",
        "annual_total_spending",
        "spending_intensity",
        "avg_monthly_total",
    ]

    TOLERANCE = 0.001

    # --- Structural Tests ---

    def test_file_exists(self):
        """Verify output Excel file was created."""
        assert os.path.exists(self.OUTPUT_PATH), f"Output file not found: {self.OUTPUT_PATH}"

    def test_valid_excel(self):
        """Verify output is a valid Excel file that can be opened."""
        wb = load_workbook(self.OUTPUT_PATH)
        assert wb is not None, "Failed to load workbook"
        wb.close()

    def test_has_answer_sheet(self):
        """Verify Answer sheet exists."""
        wb = load_workbook(self.OUTPUT_PATH)
        assert "Answer" in wb.sheetnames, "Missing required sheet: Answer"
        wb.close()

    def test_has_checks_sheet(self):
        """Verify Checks sheet exists."""
        wb = load_workbook(self.OUTPUT_PATH)
        assert "Checks" in wb.sheetnames, "Missing required sheet: Checks"
        wb.close()

    # --- Answer Sheet Structure Tests ---

    def test_answer_sheet_headers(self):
        """Verify Answer sheet has correct column headers in order."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, 6)]
        assert headers == self.EXPECTED_HEADERS, f"Headers mismatch: expected {self.EXPECTED_HEADERS}, got {headers}"
        wb.close()

    def test_answer_sheet_has_formulas(self):
        """Verify Answer sheet row 2 contains formulas referencing Checks sheet."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]
        for col in range(1, 6):
            cell_value = ws.cell(row=2, column=col).value
            assert isinstance(cell_value, str) and cell_value.startswith("="), \
                f"Column {col} should contain a formula, got: {cell_value}"
            assert "Checks!" in cell_value, \
                f"Column {col} formula should reference Checks sheet, got: {cell_value}"
        wb.close()

    def test_answer_sheet_formulas_reference_checks(self):
        """Verify Answer sheet formulas use cross-sheet references to Checks."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]

        # Check each formula contains a Checks! reference
        a2 = ws["A2"].value  # mean_max_growth_rate
        b2 = ws["B2"].value  # best_year
        c2 = ws["C2"].value  # annual_total_spending
        d2 = ws["D2"].value  # spending_intensity
        e2 = ws["E2"].value  # avg_monthly_total

        assert "Checks!" in a2, "mean_max_growth_rate should reference Checks sheet"
        assert "Checks!" in b2, "best_year should reference Checks sheet"
        assert "Checks!" in c2, "annual_total_spending should reference Checks sheet"
        assert "Checks!" in d2, "spending_intensity should reference Checks sheet"
        assert "Checks!" in e2, "avg_monthly_total should reference Checks sheet"
        wb.close()

    # --- Number Format Tests ---

    def test_decimal_format_mean_max_growth_rate(self):
        """Verify mean_max_growth_rate has 3-decimal format."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]
        fmt = ws["A2"].number_format
        assert fmt == "0.000", f"mean_max_growth_rate format should be '0.000', got '{fmt}'"
        wb.close()

    def test_decimal_format_spending_intensity(self):
        """Verify spending_intensity has 3-decimal format."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]
        fmt = ws["D2"].number_format
        assert fmt == "0.000", f"spending_intensity format should be '0.000', got '{fmt}'"
        wb.close()

    def test_currency_format_annual_total_spending(self):
        """Verify annual_total_spending has currency format."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]
        fmt = ws["C2"].number_format
        # Expected format: $#,##0;($#,##0);"-"
        assert "$" in fmt, f"annual_total_spending should have currency format with $, got '{fmt}'"
        assert '"-"' in fmt, f"Currency format should show '-' for zero, got '{fmt}'"
        wb.close()

    def test_currency_format_avg_monthly_total(self):
        """Verify avg_monthly_total has currency format."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]
        fmt = ws["E2"].number_format
        # Expected format: $#,##0;($#,##0);"-"
        assert "$" in fmt, f"avg_monthly_total should have currency format with $, got '{fmt}'"
        wb.close()

    # --- IFERROR Tests ---

    def test_iferror_in_annual_total_spending(self):
        """Verify annual_total_spending formula uses IFERROR."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]
        formula = ws["C2"].value
        assert "IFERROR" in formula.upper(), f"annual_total_spending should use IFERROR, got: {formula}"
        wb.close()

    def test_iferror_in_spending_intensity(self):
        """Verify spending_intensity formula uses IFERROR."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Answer"]
        formula = ws["D2"].value
        assert "IFERROR" in formula.upper(), f"spending_intensity should use IFERROR, got: {formula}"
        wb.close()

    # --- Checks Sheet Structure Tests ---

    def test_checks_has_pivot_table(self):
        """Verify Checks sheet has pivot table of max_growth_rate by Strain/Replicate."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Checks"]

        # Row 2 should have 'Strain' as first header
        assert ws["A2"].value == "Strain", "Pivot table should have 'Strain' header in A2"

        # Should have replicate numbers as column headers
        replicates = [ws.cell(row=2, column=i).value for i in range(2, 11)]
        assert 1 in replicates, "Pivot table should have replicate 1 in header"
        assert 9 in replicates, "Pivot table should have replicate 9 in header"
        wb.close()

    def test_checks_has_strain_rows(self):
        """Verify Checks sheet pivot table has C6706 and LuxO strains."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Checks"]

        strains = [ws.cell(row=r, column=1).value for r in range(3, 10)]
        assert "C6706" in strains, "Pivot table should have C6706 strain"
        assert "LuxO" in strains, "Pivot table should have LuxO strain"
        wb.close()

    def test_checks_has_year_table(self):
        """Verify Checks sheet has year-by-year spending table."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Checks"]

        # Find year table header
        found_year_table = False
        for row in range(1, 30):
            if ws.cell(row=row, column=1).value == "time_year":
                found_year_table = True
                # Check table headers
                assert ws.cell(row=row, column=2).value == "annual_total_spending", \
                    "Year table should have annual_total_spending column"
                assert ws.cell(row=row, column=3).value == "spending_intensity", \
                    "Year table should have spending_intensity column"
                break

        assert found_year_table, "Checks sheet should have year-by-year table with 'time_year' header"
        wb.close()

    # --- Value Tests (from Checks sheet summary metrics) ---

    def test_mean_max_growth_rate_value(self):
        """Verify mean_max_growth_rate value is correct."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Checks"]

        # Find mean_max_growth_rate in summary
        value = None
        for row in range(20, 35):
            if ws.cell(row=row, column=1).value == "mean_max_growth_rate":
                value = ws.cell(row=row, column=2).value
                break

        assert value is not None, "Could not find mean_max_growth_rate in Checks sheet"
        assert math.isclose(value, self.EXPECTED_RESULT["mean_max_growth_rate"], rel_tol=self.TOLERANCE), \
            f"mean_max_growth_rate mismatch: expected {self.EXPECTED_RESULT['mean_max_growth_rate']}, got {value}"
        wb.close()

    def test_best_year_value(self):
        """Verify best_year value is correct."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Checks"]

        # Find best_year in summary
        value = None
        for row in range(20, 35):
            if ws.cell(row=row, column=1).value == "best_year":
                value = ws.cell(row=row, column=2).value
                break

        assert value is not None, "Could not find best_year in Checks sheet"
        assert value == self.EXPECTED_RESULT["best_year"], \
            f"best_year mismatch: expected {self.EXPECTED_RESULT['best_year']}, got {value}"
        wb.close()

    def test_annual_total_spending_value(self):
        """Verify annual_total_spending value is correct."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Checks"]

        # Find best_annual_total_spending in summary
        value = None
        for row in range(20, 35):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and "annual_total_spending" in str(cell_val):
                value = ws.cell(row=row, column=2).value
                break

        assert value is not None, "Could not find annual_total_spending in Checks sheet"
        assert math.isclose(value, self.EXPECTED_RESULT["annual_total_spending"], rel_tol=self.TOLERANCE), \
            f"annual_total_spending mismatch: expected {self.EXPECTED_RESULT['annual_total_spending']}, got {value}"
        wb.close()

    def test_spending_intensity_value(self):
        """Verify spending_intensity value is correct."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Checks"]

        # Find best_spending_intensity in summary
        value = None
        for row in range(20, 35):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and "spending_intensity" in str(cell_val) and "best" in str(cell_val):
                value = ws.cell(row=row, column=2).value
                break

        assert value is not None, "Could not find spending_intensity in Checks sheet"
        assert math.isclose(value, self.EXPECTED_RESULT["spending_intensity"], rel_tol=self.TOLERANCE), \
            f"spending_intensity mismatch: expected {self.EXPECTED_RESULT['spending_intensity']}, got {value}"
        wb.close()

    def test_avg_monthly_total_is_blank(self):
        """Verify avg_monthly_total is blank (None) since 2007 not in construction.csv."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Checks"]

        # Find avg_monthly_total in summary
        value = "NOT_FOUND"
        for row in range(20, 35):
            if ws.cell(row=row, column=1).value == "avg_monthly_total":
                value = ws.cell(row=row, column=2).value
                break

        assert value != "NOT_FOUND", "Could not find avg_monthly_total in Checks sheet"
        assert value is None or value == "", \
            f"avg_monthly_total should be blank (None), got {value}"
        wb.close()

    # --- Data Integrity Tests ---

    def test_growth_rate_values_reasonable(self):
        """Verify max growth rate values are in reasonable range (0 to 1)."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Checks"]

        # Check values in pivot table (rows 3-4, columns 2-10)
        for row in range(3, 5):
            for col in range(2, 11):
                value = ws.cell(row=row, column=col).value
                if value is not None:
                    assert 0 < value < 1, \
                        f"Growth rate at ({row},{col}) = {value} should be between 0 and 1"
        wb.close()

    def test_year_table_has_multiple_years(self):
        """Verify year table has data for multiple years."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Checks"]

        # Find year table and count years
        years = []
        in_year_table = False
        for row in range(1, 30):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val == "time_year":
                in_year_table = True
                continue
            if in_year_table:
                if isinstance(cell_val, (int, float)) and 2000 <= cell_val <= 2030:
                    years.append(cell_val)
                elif cell_val and str(cell_val).isdigit():
                    year = int(cell_val)
                    if 2000 <= year <= 2030:
                        years.append(year)
                else:
                    break  # End of year table

        assert len(years) >= 5, f"Year table should have at least 5 years, got {len(years)}"
        assert 2007 in years, "Year table should include best_year 2007"
        wb.close()

    def test_year_2007_has_highest_spending_intensity(self):
        """Verify 2007 has the highest spending intensity in the year table."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb["Checks"]

        # Find year table and extract spending intensities
        year_intensities = {}
        in_year_table = False
        for row in range(1, 30):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val == "time_year":
                in_year_table = True
                continue
            if in_year_table:
                year = cell_val
                if isinstance(year, (int, float)) and 2000 <= year <= 2030:
                    intensity = ws.cell(row=row, column=3).value
                    if intensity is not None:
                        year_intensities[int(year)] = intensity
                elif year is None or (isinstance(year, str) and not year.isdigit()):
                    break

        # Find year with max intensity
        max_year = max(year_intensities, key=year_intensities.get)
        assert max_year == 2007, \
            f"Year with highest spending intensity should be 2007, got {max_year}"
        wb.close()
